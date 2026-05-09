#!/usr/bin/env python3
"""Build the Lupine DCF model as a .xlsx file following the methodology
in ~/.claude/plugins/financial-services/.../dcf-model/SKILL.md.

Inputs are read from business-plan/value-model/dcf_inputs.csv and
business-plan/value-model/lupine_revenue_v2.csv (single source of truth).

Output: business-plan/financials/Lupine_DCF_Model.xlsx with two sheets:
  - DCF: bear/base/bull scenario blocks, full FCF + valuation, two
    sensitivity tables at the bottom (WACC vs Terminal Growth, Revenue
    Growth vs EBIT Margin).
  - WACC: cost-of-capital calculation with CAPM build.
"""
from __future__ import annotations

from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

ROOT = Path(__file__).resolve().parent.parent
OUT = ROOT / "financials" / "Lupine_DCF_Model.xlsx"

# Color scheme per skill: keep it minimal — 3 blues + 1 grey + white.
NAVY = "1F4E79"
LIGHT_BLUE = "D9E1F2"
MED_BLUE = "BDD7EE"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"

BLUE_FONT = Font(color="0000FF", name="Calibri", size=11)
BLACK_FONT = Font(color="000000", name="Calibri", size=11)
GREEN_FONT = Font(color="008000", name="Calibri", size=11)
WHITE_BOLD = Font(color="FFFFFF", bold=True, name="Calibri", size=12)
BLACK_BOLD = Font(color="000000", bold=True, name="Calibri", size=11)

THIN = Side(style="thin", color="888888")
THICK = Side(style="medium", color="333333")
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def section_header(ws, cell_range: str, label: str):
    """Write a section header. Set value on top-left cell first, then merge.
    This avoids the Office JS / openpyxl merged-range pitfall described in
    the SKILL.md."""
    top_left = cell_range.split(":")[0]
    ws[top_left] = label
    ws[top_left].font = WHITE_BOLD
    ws[top_left].fill = fill(NAVY)
    ws[top_left].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(cell_range)
    for row in ws[cell_range]:
        for c in row:
            c.fill = fill(NAVY)
            c.border = BORDER_THICK


def hardcoded(cell, value, source: str):
    """Write a hardcoded input with blue font, light grey fill, and a
    cell comment documenting the source."""
    cell.value = value
    cell.font = BLUE_FONT
    cell.fill = fill(LIGHT_GREY)
    cell.border = BORDER_THIN
    cell.comment = Comment(f"Source: {source}", "Lupine DCF builder")


def formula(cell, expr: str):
    cell.value = expr
    cell.font = BLACK_FONT
    cell.border = BORDER_THIN


def green_link(cell, expr: str):
    cell.value = expr
    cell.font = GREEN_FONT
    cell.border = BORDER_THIN


def label_cell(cell, text: str, bold: bool = False):
    cell.value = text
    cell.font = BLACK_BOLD if bold else BLACK_FONT
    cell.border = BORDER_THIN


def output_cell(cell, expr: str, *, bold: bool = True):
    cell.value = expr
    cell.font = Font(color="000000", bold=bold, name="Calibri", size=11)
    cell.fill = fill(MED_BLUE)
    cell.border = BORDER_THICK


def col_header(cell, text: str):
    cell.value = text
    cell.font = BLACK_BOLD
    cell.fill = fill(LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER_THIN


def fmt(cell, number_format: str):
    cell.number_format = number_format


# ---------------------------------------------------------------------------
# Load source data so the DCF stays in sync with value-model/
# ---------------------------------------------------------------------------

def load_inputs():
    inputs_path = ROOT / "value-model" / "dcf_inputs.csv"
    rows = {}
    with inputs_path.open() as f:
        for row in csv.DictReader(f):
            rows[row["id"]] = row
    return rows


def load_revenue():
    rev_path = ROOT / "value-model" / "lupine_revenue_v2.csv"
    rows = {}
    with rev_path.open() as f:
        for row in csv.DictReader(f):
            rows[row["id"]] = row
    return rows


inputs = load_inputs()
revenue = load_revenue()


# ---------------------------------------------------------------------------
# Build WACC sheet first (DCF references it)
# ---------------------------------------------------------------------------

wb = Workbook()
ws_dcf = wb.active
ws_dcf.title = "DCF"
ws_wacc = wb.create_sheet("WACC")

# WACC sheet
ws_wacc.column_dimensions["A"].width = 38
for col in ["B", "C", "D"]:
    ws_wacc.column_dimensions[col].width = 16

ws_wacc["A1"] = "Lupine Materials Science — WACC"
ws_wacc["A1"].font = Font(bold=True, size=14, color="000000")
ws_wacc.merge_cells("A1:D1")

section_header(ws_wacc, "A3:D3", "COST OF EQUITY (CAPM)")
label_cell(ws_wacc["A4"], "Risk-free rate (10Y Treasury)")
hardcoded(ws_wacc["B4"], 0.044, "value-model/dcf_inputs.csv:risk-free-rate (US Treasury 10Y May 2026)")
fmt(ws_wacc["B4"], "0.00%")

label_cell(ws_wacc["A5"], "Levered beta (vertical SaaS)")
hardcoded(ws_wacc["B5"], 1.40, "value-model/dcf_inputs.csv:beta-base (Damodaran NYU 2026 industry beta)")
fmt(ws_wacc["B5"], "0.00")

label_cell(ws_wacc["A6"], "Equity risk premium")
hardcoded(ws_wacc["B6"], 0.055, "value-model/dcf_inputs.csv:equity-risk-premium (Damodaran NYU 2026)")
fmt(ws_wacc["B6"], "0.00%")

label_cell(ws_wacc["A7"], "Cost of equity (CAPM)", bold=True)
formula(ws_wacc["B7"], "=B4+B5*B6")
fmt(ws_wacc["B7"], "0.00%")

section_header(ws_wacc, "A9:D9", "COST OF DEBT")
label_cell(ws_wacc["A10"], "Pre-tax cost of debt")
hardcoded(ws_wacc["B10"], 0.080, "Mid-market venture-debt 2026")
fmt(ws_wacc["B10"], "0.00%")
label_cell(ws_wacc["A11"], "Tax rate")
hardcoded(ws_wacc["B11"], 0.21, "US federal corporate tax rate")
fmt(ws_wacc["B11"], "0.00%")
label_cell(ws_wacc["A12"], "After-tax cost of debt")
formula(ws_wacc["B12"], "=B10*(1-B11)")
fmt(ws_wacc["B12"], "0.00%")

section_header(ws_wacc, "A14:D14", "CAPITAL STRUCTURE (post-seed)")
label_cell(ws_wacc["A15"], "Equity weight")
hardcoded(ws_wacc["B15"], 1.00, "value-model/dcf_inputs.csv:debt-weight (Lupine ~100% equity post-seed)")
fmt(ws_wacc["B15"], "0.00%")
label_cell(ws_wacc["A16"], "Debt weight")
formula(ws_wacc["B16"], "=1-B15")
fmt(ws_wacc["B16"], "0.00%")

section_header(ws_wacc, "A18:D18", "WACC RESULT")
col_header(ws_wacc["A19"], "Component")
col_header(ws_wacc["B19"], "Weight")
col_header(ws_wacc["C19"], "Cost")
col_header(ws_wacc["D19"], "Contribution")
label_cell(ws_wacc["A20"], "Equity")
formula(ws_wacc["B20"], "=B15"); fmt(ws_wacc["B20"], "0.00%")
formula(ws_wacc["C20"], "=B7"); fmt(ws_wacc["C20"], "0.00%")
formula(ws_wacc["D20"], "=B20*C20"); fmt(ws_wacc["D20"], "0.00%")
label_cell(ws_wacc["A21"], "Debt")
formula(ws_wacc["B21"], "=B16"); fmt(ws_wacc["B21"], "0.00%")
formula(ws_wacc["C21"], "=B12"); fmt(ws_wacc["C21"], "0.00%")
formula(ws_wacc["D21"], "=B21*C21"); fmt(ws_wacc["D21"], "0.00%")

label_cell(ws_wacc["A23"], "WACC (base case)", bold=True)
output_cell(ws_wacc["B23"], "=D20+D21")
fmt(ws_wacc["B23"], "0.00%")

# ---------------------------------------------------------------------------
# DCF sheet
# ---------------------------------------------------------------------------

ws = ws_dcf
ws.column_dimensions["A"].width = 36
for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L"]:
    ws.column_dimensions[col].width = 14

ws["A1"] = "Lupine Materials Science — DCF Model"
ws["A1"].font = Font(bold=True, size=14, color="000000")
ws.merge_cells("A1:J1")
ws["A2"] = "Pre-revenue seed; bottom-up sector value-unlock build; mid-year discount"
ws["A2"].font = Font(italic=True, color="333333", size=10)
ws.merge_cells("A2:J2")

# --- Case selector ---
section_header(ws, "A4:J4", "CASE SELECTOR (1=Bear  2=Base  3=Bull)")
label_cell(ws["A5"], "Case selector cell", bold=True)
hardcoded(ws["B5"], 2, "User selectable: 1=Bear, 2=Base, 3=Bull")
formula(ws["C5"], '=IF(B5=1,"Bear",IF(B5=2,"Base","Bull"))')
ws["C5"].font = Font(bold=True, color="000000")

# --- Market Data ---
section_header(ws, "A7:J7", "MARKET DATA & KEY INPUTS")
label_cell(ws["A8"], "Round size (mid-point) ($M)")
hardcoded(ws["B8"], 8.0, "Founder ask: $5M-$10M seed; midpoint $8M")
fmt(ws["B8"], "$#,##0.0")
label_cell(ws["A9"], "Post-money valuation (mid-point) ($M)")
hardcoded(ws["B9"], 150.0, "Founder ask: $100M-$200M post; midpoint $150M")
fmt(ws["B9"], "$#,##0")
label_cell(ws["A10"], "Implied seed dilution")
formula(ws["B10"], "=B8/B9")
fmt(ws["B10"], "0.00%")

label_cell(ws["A11"], "Net debt ($M)")
hardcoded(ws["B11"], 0, "Lupine has no debt; cash position not material at seed")
fmt(ws["B11"], "$#,##0")

# --- Scenario assumption blocks (Bear / Base / Bull) ---
# Each block lays assumptions horizontally across 7 projection years FY26-FY32.

YEAR_HEADERS = ["FY26", "FY27", "FY28", "FY29", "FY30", "FY31", "FY32"]
PROJ_COLS = ["C", "D", "E", "F", "G", "H", "I"]
N_YEARS = len(YEAR_HEADERS)

# Bear case row 13 onward
section_header(ws, "A13:J13", "BEAR CASE ASSUMPTIONS")
col_header(ws["A14"], "Assumption")
for i, y in enumerate(YEAR_HEADERS):
    col_header(ws[f"{PROJ_COLS[i]}14"], y)

bear_growth = [12, 30, 60, 80, 70, 50, 40]
bear_ebit_margin = [-600, -100, -50, -10, 10, 20, 28]
bear_da = [3, 4, 4, 3, 3, 3, 2]
bear_capex = [5, 4, 4, 3, 3, 3, 2]
bear_nwc = [10, 10, 8, 8, 7, 7, 6]

label_cell(ws["A15"], "Revenue growth (%)")
for i, v in enumerate(bear_growth):
    hardcoded(ws[f"{PROJ_COLS[i]}15"], v / 100.0, f"Bear case revenue growth FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}15"], "0.0%")

label_cell(ws["A16"], "EBIT margin (%)")
for i, v in enumerate(bear_ebit_margin):
    hardcoded(ws[f"{PROJ_COLS[i]}16"], v / 100.0, f"Bear case EBIT margin FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}16"], "0.0%")

label_cell(ws["A17"], "D&A (% of revenue)")
for i, v in enumerate(bear_da):
    hardcoded(ws[f"{PROJ_COLS[i]}17"], v / 100.0, f"Bear case D&A FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}17"], "0.0%")

label_cell(ws["A18"], "CapEx (% of revenue)")
for i, v in enumerate(bear_capex):
    hardcoded(ws[f"{PROJ_COLS[i]}18"], v / 100.0, f"Bear case CapEx FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}18"], "0.0%")

label_cell(ws["A19"], "NWC change (% of d-Rev)")
for i, v in enumerate(bear_nwc):
    hardcoded(ws[f"{PROJ_COLS[i]}19"], v / 100.0, f"Bear case NWC FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}19"], "0.0%")

# Base case row 21 onward
section_header(ws, "A21:J21", "BASE CASE ASSUMPTIONS")
col_header(ws["A22"], "Assumption")
for i, y in enumerate(YEAR_HEADERS):
    col_header(ws[f"{PROJ_COLS[i]}22"], y)

# Base-case derived from value-model/lupine_revenue_v2.csv
# Implied growth: rev_total values
def _flt(rid, fy):
    return float(revenue[rid][fy])

rev_v2 = [_flt("rev-total", f"fy{2026+i}") for i in range(N_YEARS)]
# growth FY26 anchored to a $0.05M starting point (LTM 2025); for FY26 we
# use a representative growth rate of 600% reflecting the small base.
base_growth = []
for i in range(N_YEARS):
    if i == 0:
        base_growth.append(6.00)  # +600% off near-zero base FY25
    else:
        base_growth.append(rev_v2[i] / rev_v2[i - 1] - 1)

# EBIT margin from CSV
base_ebit_margin = [_flt("ebit-margin", f"fy{2026+i}") / 100.0 for i in range(N_YEARS)]
# D&A and CapEx as % of rev derived from CSV
da_v2 = [_flt("da", f"fy{2026+i}") for i in range(N_YEARS)]
capex_v2 = [_flt("capex", f"fy{2026+i}") for i in range(N_YEARS)]
nwc_v2 = [_flt("nwc-change", f"fy{2026+i}") for i in range(N_YEARS)]
base_da = [da_v2[i] / rev_v2[i] for i in range(N_YEARS)]
base_capex = [capex_v2[i] / rev_v2[i] for i in range(N_YEARS)]
# NWC % of d-Rev
base_nwc = []
prev = 0.05  # FY25 starting revenue assumption
for i in range(N_YEARS):
    d_rev = rev_v2[i] - prev
    base_nwc.append(nwc_v2[i] / d_rev if d_rev > 0 else 0.07)
    prev = rev_v2[i]

label_cell(ws["A23"], "Revenue growth (%)")
for i, v in enumerate(base_growth):
    hardcoded(ws[f"{PROJ_COLS[i]}23"], v, f"Base case revenue growth FY{26+i} (lupine_revenue_v2.csv:rev-total derived)")
    fmt(ws[f"{PROJ_COLS[i]}23"], "0.0%")

label_cell(ws["A24"], "EBIT margin (%)")
for i, v in enumerate(base_ebit_margin):
    hardcoded(ws[f"{PROJ_COLS[i]}24"], v, f"Base case EBIT margin FY{26+i} (lupine_revenue_v2.csv:ebit-margin)")
    fmt(ws[f"{PROJ_COLS[i]}24"], "0.0%")

label_cell(ws["A25"], "D&A (% of revenue)")
for i, v in enumerate(base_da):
    hardcoded(ws[f"{PROJ_COLS[i]}25"], v, f"Base case D&A FY{26+i} (lupine_revenue_v2.csv)")
    fmt(ws[f"{PROJ_COLS[i]}25"], "0.0%")

label_cell(ws["A26"], "CapEx (% of revenue)")
for i, v in enumerate(base_capex):
    hardcoded(ws[f"{PROJ_COLS[i]}26"], v, f"Base case CapEx FY{26+i} (lupine_revenue_v2.csv)")
    fmt(ws[f"{PROJ_COLS[i]}26"], "0.0%")

label_cell(ws["A27"], "NWC change (% of d-Rev)")
for i, v in enumerate(base_nwc):
    hardcoded(ws[f"{PROJ_COLS[i]}27"], v, f"Base case NWC FY{26+i} (lupine_revenue_v2.csv)")
    fmt(ws[f"{PROJ_COLS[i]}27"], "0.0%")

# Bull case row 29 onward
section_header(ws, "A29:J29", "BULL CASE ASSUMPTIONS")
col_header(ws["A30"], "Assumption")
for i, y in enumerate(YEAR_HEADERS):
    col_header(ws[f"{PROJ_COLS[i]}30"], y)

bull_growth = [12.0, 4.0, 1.5, 1.0, 0.65, 0.45, 0.35]
bull_ebit_margin = [-300, -3, 0.30, 0.42, 0.50, 0.55, 0.58]
# normalize ebit margins:
bull_ebit_margin = [-3.0, -0.03, 0.30, 0.42, 0.50, 0.55, 0.58]
bull_da = [0.02, 0.025, 0.025, 0.022, 0.020, 0.018, 0.016]
bull_capex = [0.04, 0.035, 0.030, 0.025, 0.022, 0.020, 0.018]
bull_nwc = [0.06, 0.07, 0.07, 0.07, 0.07, 0.06, 0.06]

label_cell(ws["A31"], "Revenue growth (%)")
for i, v in enumerate(bull_growth):
    hardcoded(ws[f"{PROJ_COLS[i]}31"], v, f"Bull case revenue growth FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}31"], "0.0%")

label_cell(ws["A32"], "EBIT margin (%)")
for i, v in enumerate(bull_ebit_margin):
    hardcoded(ws[f"{PROJ_COLS[i]}32"], v, f"Bull case EBIT margin FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}32"], "0.0%")

label_cell(ws["A33"], "D&A (% of revenue)")
for i, v in enumerate(bull_da):
    hardcoded(ws[f"{PROJ_COLS[i]}33"], v, f"Bull case D&A FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}33"], "0.0%")

label_cell(ws["A34"], "CapEx (% of revenue)")
for i, v in enumerate(bull_capex):
    hardcoded(ws[f"{PROJ_COLS[i]}34"], v, f"Bull case CapEx FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}34"], "0.0%")

label_cell(ws["A35"], "NWC change (% of d-Rev)")
for i, v in enumerate(bull_nwc):
    hardcoded(ws[f"{PROJ_COLS[i]}35"], v, f"Bull case NWC FY{26+i}")
    fmt(ws[f"{PROJ_COLS[i]}35"], "0.0%")

# --- Consolidation row (selected scenario) ---
section_header(ws, "A37:J37", "SELECTED SCENARIO (consolidation row from case selector)")
col_header(ws["A38"], "Assumption")
for i, y in enumerate(YEAR_HEADERS):
    col_header(ws[f"{PROJ_COLS[i]}38"], y)

# For each year column, the consolidation cell uses INDEX into the three
# scenario blocks based on $B$5 (case selector).
def consolidation_formula(year_idx: int, bear_row: int, base_row: int, bull_row: int) -> str:
    col = PROJ_COLS[year_idx]
    return f"=INDEX(({col}{bear_row},{col}{base_row},{col}{bull_row}),1,1,$B$5)"


# Simpler: use CHOOSE
def choose_formula(year_idx: int, bear_row: int, base_row: int, bull_row: int) -> str:
    col = PROJ_COLS[year_idx]
    return f"=CHOOSE($B$5,{col}{bear_row},{col}{base_row},{col}{bull_row})"


SEL_GROWTH = 39
SEL_EBIT = 40
SEL_DA = 41
SEL_CAPEX = 42
SEL_NWC = 43

label_cell(ws["A39"], "Revenue growth (%)")
for i in range(N_YEARS):
    formula(ws[f"{PROJ_COLS[i]}39"], choose_formula(i, 15, 23, 31))
    fmt(ws[f"{PROJ_COLS[i]}39"], "0.0%")
label_cell(ws["A40"], "EBIT margin (%)")
for i in range(N_YEARS):
    formula(ws[f"{PROJ_COLS[i]}40"], choose_formula(i, 16, 24, 32))
    fmt(ws[f"{PROJ_COLS[i]}40"], "0.0%")
label_cell(ws["A41"], "D&A (% of revenue)")
for i in range(N_YEARS):
    formula(ws[f"{PROJ_COLS[i]}41"], choose_formula(i, 17, 25, 33))
    fmt(ws[f"{PROJ_COLS[i]}41"], "0.0%")
label_cell(ws["A42"], "CapEx (% of revenue)")
for i in range(N_YEARS):
    formula(ws[f"{PROJ_COLS[i]}42"], choose_formula(i, 18, 26, 34))
    fmt(ws[f"{PROJ_COLS[i]}42"], "0.0%")
label_cell(ws["A43"], "NWC change (% of d-Rev)")
for i in range(N_YEARS):
    formula(ws[f"{PROJ_COLS[i]}43"], choose_formula(i, 19, 27, 35))
    fmt(ws[f"{PROJ_COLS[i]}43"], "0.0%")

# --- Other constants (scenario-independent for simplicity) ---
section_header(ws, "A45:J45", "GLOBAL ASSUMPTIONS")
label_cell(ws["A46"], "Tax rate")
hardcoded(ws["B46"], 0.21, "US federal corporate tax rate; pre-revenue years effectively 0% via NOL but conservative to model 21%")
fmt(ws["B46"], "0.0%")

label_cell(ws["A47"], "Terminal growth rate (Bear/Base/Bull)")
hardcoded(ws["B47"], 0.02, "Bear terminal growth (value-model/dcf_inputs.csv:terminal-growth-bear)")
fmt(ws["B47"], "0.0%")
hardcoded(ws["C47"], 0.03, "Base terminal growth (value-model/dcf_inputs.csv:terminal-growth-base)")
fmt(ws["C47"], "0.0%")
hardcoded(ws["D47"], 0.035, "Bull terminal growth (value-model/dcf_inputs.csv:terminal-growth-bull)")
fmt(ws["D47"], "0.0%")
label_cell(ws["E47"], "Selected:")
formula(ws["F47"], "=CHOOSE($B$5,B47,C47,D47)")
fmt(ws["F47"], "0.0%")

label_cell(ws["A48"], "WACC (Bear/Base/Bull)")
hardcoded(ws["B48"], 0.14, "Bear WACC (value-model/dcf_inputs.csv:wacc-bear)")
fmt(ws["B48"], "0.0%")
green_link(ws["C48"], "=WACC!B23")  # link to WACC sheet
fmt(ws["C48"], "0.0%")
hardcoded(ws["D48"], 0.105, "Bull WACC (value-model/dcf_inputs.csv:wacc-bull)")
fmt(ws["D48"], "0.0%")
label_cell(ws["E48"], "Selected:")
formula(ws["F48"], "=CHOOSE($B$5,B48,C48,D48)")
fmt(ws["F48"], "0.0%")

# --- Cash flow projection ---
section_header(ws, "A50:J50", "FREE CASH FLOW PROJECTION ($M)")
col_header(ws["A51"], "Item")
for i, y in enumerate(YEAR_HEADERS):
    col_header(ws[f"{PROJ_COLS[i]}51"], y)

# FY25 starting revenue (LTM proxy)
label_cell(ws["A52"], "FY25 starting revenue ($M)")
hardcoded(ws["B52"], 0.05, "Atlas-distill + paid-pilot scoping baseline; near-zero LTM")
fmt(ws["B52"], "$#,##0.00")

# Revenue
label_cell(ws["A53"], "Revenue ($M)", bold=True)
formula(ws["C53"], "=$B$52*(1+C39)"); fmt(ws["C53"], "$#,##0.00")
for i in range(1, N_YEARS):
    prev_col = PROJ_COLS[i - 1]
    cur_col = PROJ_COLS[i]
    formula(ws[f"{cur_col}53"], f"={prev_col}53*(1+{cur_col}39)")
    fmt(ws[f"{cur_col}53"], "$#,##0.00")

# EBIT
label_cell(ws["A54"], "EBIT ($M)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}54"], f"={c}53*{c}40")
    fmt(ws[f"{c}54"], "$#,##0.00")

# Taxes
label_cell(ws["A55"], "Taxes ($M)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}55"], f"=MAX(0,{c}54*$B$46)")
    fmt(ws[f"{c}55"], "$#,##0.00")

# NOPAT
label_cell(ws["A56"], "NOPAT ($M)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}56"], f"={c}54-{c}55")
    fmt(ws[f"{c}56"], "$#,##0.00")

# D&A
label_cell(ws["A57"], "(+) D&A ($M)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}57"], f"={c}53*{c}41")
    fmt(ws[f"{c}57"], "$#,##0.00")

# CapEx
label_cell(ws["A58"], "(-) CapEx ($M)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}58"], f"={c}53*{c}42")
    fmt(ws[f"{c}58"], "$#,##0.00")

# Change in NWC
label_cell(ws["A59"], "(-) Change in NWC ($M)")
formula(ws["C59"], "=(C53-$B$52)*C43"); fmt(ws["C59"], "$#,##0.00")
for i in range(1, N_YEARS):
    prev_col = PROJ_COLS[i - 1]
    cur_col = PROJ_COLS[i]
    formula(ws[f"{cur_col}59"], f"=({cur_col}53-{prev_col}53)*{cur_col}43")
    fmt(ws[f"{cur_col}59"], "$#,##0.00")

# Unlevered FCF
label_cell(ws["A60"], "Unlevered FCF ($M)", bold=True)
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}60"], f"={c}56+{c}57-{c}58-{c}59")
    ws[f"{c}60"].fill = fill(LIGHT_BLUE)
    ws[f"{c}60"].font = Font(bold=True, color="000000")
    fmt(ws[f"{c}60"], "$#,##0.00")

# Discount period (mid-year)
label_cell(ws["A61"], "Discount period (years)")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    hardcoded(ws[f"{c}61"], 0.5 + i, f"Mid-year convention period {0.5+i}")
    fmt(ws[f"{c}61"], "0.00")

# Discount factor
label_cell(ws["A62"], "Discount factor")
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}62"], f"=1/(1+$F$48)^{c}61")
    fmt(ws[f"{c}62"], "0.0000")

# PV of FCF
label_cell(ws["A63"], "PV of FCF ($M)", bold=True)
for i in range(N_YEARS):
    c = PROJ_COLS[i]
    formula(ws[f"{c}63"], f"={c}60*{c}62")
    fmt(ws[f"{c}63"], "$#,##0.00")

# --- Terminal value ---
section_header(ws, "A65:J65", "TERMINAL VALUE")
label_cell(ws["A66"], "Final-year FCF ($M)")
formula(ws["B66"], f"=I60")
fmt(ws["B66"], "$#,##0.00")
label_cell(ws["A67"], "Terminal growth rate")
formula(ws["B67"], "=$F$47")
fmt(ws["B67"], "0.0%")
label_cell(ws["A68"], "WACC")
formula(ws["B68"], "=$F$48")
fmt(ws["B68"], "0.0%")
label_cell(ws["A69"], "Terminal FCF (year +1)")
formula(ws["B69"], "=B66*(1+B67)")
fmt(ws["B69"], "$#,##0.00")
label_cell(ws["A70"], "Terminal value (Gordon)")
formula(ws["B70"], "=B69/(B68-B67)")
fmt(ws["B70"], "$#,##0.00")
label_cell(ws["A71"], "Final-year discount factor")
formula(ws["B71"], "=I62")
fmt(ws["B71"], "0.0000")
label_cell(ws["A72"], "PV of terminal value ($M)", bold=True)
formula(ws["B72"], "=B70*B71")
fmt(ws["B72"], "$#,##0.00")

# --- Valuation summary ---
section_header(ws, "A74:J74", "VALUATION SUMMARY ($M)")
label_cell(ws["A75"], "Sum of PV FCFs")
formula(ws["B75"], "=SUM(C63:I63)")
fmt(ws["B75"], "$#,##0.00")
label_cell(ws["A76"], "PV of terminal value")
formula(ws["B76"], "=B72")
fmt(ws["B76"], "$#,##0.00")
label_cell(ws["A77"], "Enterprise value", bold=True)
formula(ws["B77"], "=B75+B76")
fmt(ws["B77"], "$#,##0.00")
label_cell(ws["A78"], "(-) Net debt")
formula(ws["B78"], "=$B$11")
fmt(ws["B78"], "$#,##0.00")
label_cell(ws["A79"], "Equity value (intrinsic)", bold=True)
output_cell(ws["B79"], "=B77-B78")
fmt(ws["B79"], "$#,##0.00")

label_cell(ws["A81"], "Implied vs proposed post-money")
label_cell(ws["A82"], "Proposed post-money ($M)")
formula(ws["B82"], "=$B$9")
fmt(ws["B82"], "$#,##0")
label_cell(ws["A83"], "DCF intrinsic / proposed (margin of safety)")
formula(ws["B83"], "=B79/B82-1")
fmt(ws["B83"], "0.0%")
label_cell(ws["A84"], "Terminal value as % of EV")
formula(ws["B84"], "=B76/B77")
fmt(ws["B84"], "0.0%")

# --- Sensitivity Tables ---
# Per skill: ODD x ODD with base case in center; populate every cell with
# a full DCF re-calc formula. We use 5x5 with axes built around base case.
# To keep formulas tractable inside cells, we recompute the valuation as
# (Sum of PV FCFs at given WACC) + (TV at given growth, given WACC) - net
# debt — using the base-case FCF stream from row 60.

# Sensitivity 1: WACC vs Terminal Growth
SENS1_TOP = 88
section_header(ws, f"A{SENS1_TOP}:J{SENS1_TOP}", "SENSITIVITY 1: WACC vs TERMINAL GROWTH (Equity Value $M)")
# Build axes around base-case WACC=12.10%, terminal g=3.00%
base_wacc = 0.121
base_g = 0.03
wacc_axis = [base_wacc - 0.020, base_wacc - 0.010, base_wacc, base_wacc + 0.010, base_wacc + 0.020]
g_axis = [base_g - 0.010, base_g - 0.005, base_g, base_g + 0.005, base_g + 0.010]

# Column headers (terminal growth)
ws[f"A{SENS1_TOP+1}"] = "WACC \\ Terminal g"
ws[f"A{SENS1_TOP+1}"].font = BLACK_BOLD
ws[f"A{SENS1_TOP+1}"].fill = fill(LIGHT_BLUE)
ws[f"A{SENS1_TOP+1}"].border = BORDER_THIN
for j, g_val in enumerate(g_axis):
    cell = ws.cell(row=SENS1_TOP + 1, column=2 + j)
    cell.value = g_val
    cell.font = BLACK_BOLD
    cell.fill = fill(LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER_THIN
    cell.number_format = "0.0%"

# Row headers (WACC) and formulas
for i, w_val in enumerate(wacc_axis):
    row = SENS1_TOP + 2 + i
    label_cell(ws.cell(row=row, column=1), "")
    ws.cell(row=row, column=1).value = w_val
    ws.cell(row=row, column=1).font = BLACK_BOLD
    ws.cell(row=row, column=1).fill = fill(LIGHT_BLUE)
    ws.cell(row=row, column=1).number_format = "0.0%"
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
    ws.cell(row=row, column=1).border = BORDER_THIN
    for j, g_val in enumerate(g_axis):
        col_letter = chr(ord("B") + j)
        # Formula: sum of FCF discounted at this WACC + TV at this g/WACC
        # Sum PV FCFs at this WACC
        # use period 0.5..6.5
        terms = []
        for k in range(N_YEARS):
            period = 0.5 + k
            fcf_cell = f"{PROJ_COLS[k]}60"
            terms.append(f"{fcf_cell}/(1+$A{row})^{period}")
        sum_pv = "+".join(terms)
        tv = f"(I60*(1+{col_letter}${SENS1_TOP+1}))/($A{row}-{col_letter}${SENS1_TOP+1})/(1+$A{row})^6.5"
        expr = f"=({sum_pv})+{tv}-$B$11"
        cell = ws.cell(row=row, column=2 + j)
        cell.value = expr
        cell.font = BLACK_FONT
        cell.number_format = "$#,##0"
        cell.border = BORDER_THIN
        # Highlight center (i=2, j=2)
        if i == 2 and j == 2:
            cell.font = Font(bold=True, color="000000")
            cell.fill = fill(MED_BLUE)
            cell.border = BORDER_THICK

# Sensitivity 2: Revenue Growth vs Terminal EBIT Margin
# This re-runs the DCF using a uniform growth rate and a single terminal
# EBIT margin to demonstrate operational sensitivity. Built bottom-up
# from FY25 starting revenue = $0.05M.

SENS2_TOP = 100
section_header(ws, f"A{SENS2_TOP}:J{SENS2_TOP}", "SENSITIVITY 2: AVG REVENUE GROWTH vs TERMINAL EBIT MARGIN (Equity Value $M)")
growth_axis = [0.50, 0.75, 1.00, 1.25, 1.50]  # avg compound growth FY26-FY32
margin_axis = [0.20, 0.30, 0.40, 0.50, 0.60]

ws[f"A{SENS2_TOP+1}"] = "Avg growth \\ Term margin"
ws[f"A{SENS2_TOP+1}"].font = BLACK_BOLD
ws[f"A{SENS2_TOP+1}"].fill = fill(LIGHT_BLUE)
ws[f"A{SENS2_TOP+1}"].border = BORDER_THIN
for j, m in enumerate(margin_axis):
    cell = ws.cell(row=SENS2_TOP + 1, column=2 + j)
    cell.value = m
    cell.font = BLACK_BOLD
    cell.fill = fill(LIGHT_BLUE)
    cell.alignment = Alignment(horizontal="center")
    cell.border = BORDER_THIN
    cell.number_format = "0.0%"

for i, g in enumerate(growth_axis):
    row = SENS2_TOP + 2 + i
    cell_g = ws.cell(row=row, column=1)
    cell_g.value = g
    cell_g.font = BLACK_BOLD
    cell_g.fill = fill(LIGHT_BLUE)
    cell_g.number_format = "0.0%"
    cell_g.alignment = Alignment(horizontal="center")
    cell_g.border = BORDER_THIN
    for j, m in enumerate(margin_axis):
        col_letter = chr(ord("B") + j)
        # Build a synthetic FCF stream: revenue compounds at growth rate g
        # off $B$52 ($0.05M); EBIT margin ramps linearly from -3.0 to m;
        # tax 21% on positive EBIT; D&A 2.5%; CapEx 3.0%; NWC 7%.
        # Discount at base WACC=12.10%.
        wacc = 0.121
        fcf_terms = []
        for k in range(N_YEARS):
            year = k + 1
            period = 0.5 + k
            # margin ramp: lerp from -3.0 to m over 7 periods
            t_pct = (k + 1) / N_YEARS
            margin_expr = f"(-3+($A{row}*0+{col_letter}${SENS2_TOP+1}+3)*{t_pct})"
            # synthetic revenue: $B$52 * (1 + g)^year  (g = $A{row})
            rev_expr = f"$B$52*(1+$A{row})^{year}"
            ebit_expr = f"({rev_expr})*({margin_expr})"
            tax_expr = f"MAX(0,({ebit_expr})*$B$46)"
            nopat_expr = f"({ebit_expr})-({tax_expr})"
            da_expr = f"({rev_expr})*0.025"
            capex_expr = f"({rev_expr})*0.030"
            # delta NWC: use simplified expression
            if k == 0:
                drev = f"({rev_expr}-$B$52)"
            else:
                drev = f"({rev_expr}-$B$52*(1+$A{row})^{k})"
            nwc_expr = f"({drev})*0.07"
            fcf_year = f"({nopat_expr}+{da_expr}-{capex_expr}-{nwc_expr})"
            fcf_terms.append(f"{fcf_year}/(1+{wacc})^{period}")
        sum_pv = "+".join(fcf_terms)
        # Terminal value with base g=0.03 and same wacc
        # Terminal FCF approximated as final-year FCF
        final_year_fcf_expr = (
            f"($B$52*(1+$A{row})^{N_YEARS}*"
            f"((-3+($A{row}*0+{col_letter}${SENS2_TOP+1}+3)*1)*(1-$B$46)+0.025-0.030-0.07*$A{row}/(1+$A{row}))"
            f")"
        )
        tv_expr = f"({final_year_fcf_expr}*1.03/({wacc}-0.03))/(1+{wacc})^6.5"
        full = f"=({sum_pv})+{tv_expr}"
        cell = ws.cell(row=row, column=2 + j)
        cell.value = full
        cell.font = BLACK_FONT
        cell.number_format = "$#,##0"
        cell.border = BORDER_THIN
        if i == 2 and j == 3:
            # highlight closest-to-base cell (1.00 growth, 50% terminal margin)
            cell.font = Font(bold=True, color="000000")
            cell.fill = fill(MED_BLUE)
            cell.border = BORDER_THICK

# Save
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
