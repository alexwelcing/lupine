#!/usr/bin/env python3
"""Build the Lupine comparable-companies analysis as a .xlsx file
following the methodology in
~/.claude/plugins/financial-services/.../comps-analysis/SKILL.md.

Inputs: business-plan/value-model/comparable_companies_v2.csv
Output: business-plan/financials/Lupine_Comps_Analysis.xlsx
"""
from __future__ import annotations

from pathlib import Path
import csv

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "value-model" / "comparable_companies_v2.csv"
OUT = ROOT / "financials" / "Lupine_Comps_Analysis.xlsx"

NAVY = "1F4E79"
LIGHT_BLUE = "D9E1F2"
MED_BLUE = "BDD7EE"
LIGHT_GREY = "F2F2F2"
THIN = Side(style="thin", color="888888")
THICK = Side(style="medium", color="333333")
BORDER_THICK = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def fill(c):
    return PatternFill("solid", fgColor=c)


with SRC.open() as f:
    rows = list(csv.DictReader(f))

wb = Workbook()
ws = wb.active
ws.title = "Comps"

ws.column_dimensions["A"].width = 30
for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
    ws.column_dimensions[col].width = 14

# Title
ws["A1"] = "Lupine Materials Science — Comparable Companies"
ws["A1"].font = Font(bold=True, size=14)
ws.merge_cells("A1:K1")
ws["A2"] = "Engineering simulation + AI-for-science + materials informatics group"
ws["A2"].font = Font(italic=True, color="333333", size=10)
ws.merge_cells("A2:K2")

# Header
header_row = 4
headers = [
    "Company", "Ticker", "Sector", "Market Cap ($M)", "Revenue LTM ($M)",
    "EV ($M)", "EV/Revenue", "EV/EBITDA", "Gross Margin", "Rev Growth", "Tier",
]
for i, h in enumerate(headers):
    c = ws.cell(row=header_row, column=1 + i)
    c.value = h
    c.font = Font(bold=True, color="FFFFFF")
    c.fill = fill(NAVY)
    c.alignment = Alignment(horizontal="center")
    c.border = BORDER_THICK

# Data rows
def num(v):
    if v is None or v == "" or v == "n/a" or v == "negative":
        return None
    try:
        return float(v)
    except ValueError:
        return None


public_company_rows = [r for r in rows if not r["id"].startswith("median-")]
median_rows = [r for r in rows if r["id"].startswith("median-")]

cur = header_row + 1
for r in public_company_rows:
    mc = num(r.get("market_cap_usd_m"))
    rev = num(r.get("revenue_ltm_usd_m"))
    ev = num(r.get("ev_usd_m"))
    ev_rev = num(r.get("ev_revenue_x"))
    ev_ebitda = num(r.get("ev_ebitda_x"))
    gm = num(r.get("gross_margin_pct"))
    growth = num(r.get("rev_growth_pct"))
    cells = [
        r["company"], r["ticker"], r["sector"], mc, rev, ev, ev_rev, ev_ebitda,
        gm, growth, r["tier"],
    ]
    for i, val in enumerate(cells):
        c = ws.cell(row=cur, column=1 + i)
        c.value = val
        c.border = BORDER_THIN
        if i in (3, 4, 5):
            c.number_format = "$#,##0"
        elif i in (6, 7):
            c.number_format = "0.0\"x\""
        elif i in (8, 9):
            c.number_format = "0.0%"
            if val is not None:
                c.value = val / 100.0
        if r["company"].startswith("Schrodinger software segment"):
            c.fill = fill(MED_BLUE)
            c.font = Font(bold=True, color="000000")
        # source comment on the company name cell
        if i == 0 and r.get("source_url"):
            c.comment = Comment(
                f"Source: {r['source_url']}\nNotes: {r.get('notes', '')}",
                "Lupine Comps builder",
            )
    cur += 1

# --- Statistical row block ---
cur += 1
ws.cell(row=cur, column=1).value = "STATISTICS (public-company group, ex AI-bio outliers)"
ws.cell(row=cur, column=1).font = Font(bold=True, color="FFFFFF")
ws.cell(row=cur, column=1).fill = fill(NAVY)
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=11)
cur += 1

# Identify the row range for the public companies that we want to summarize
# Use the simulation comps (Synopsys, Cadence, ANSYS, Altair, Veeva)
public_simulation_ids = ["synopsys", "cadence", "ansys-pre", "altair", "veeva"]
public_rows_for_stats = []
for i, r in enumerate(public_company_rows):
    if r["id"] in public_simulation_ids:
        public_rows_for_stats.append(header_row + 1 + i)

# Write median formulas for the simulation comp set
ws.cell(row=cur, column=1).value = "Median (simulation comps)"
ws.cell(row=cur, column=1).font = Font(bold=True)
ws.cell(row=cur, column=1).border = BORDER_THIN
for col_idx, col_name in [(7, "EV/Revenue"), (8, "EV/EBITDA"), (9, "Gross Margin"), (10, "Rev Growth")]:
    refs = ",".join(f"{chr(ord('A')+col_idx-1)}{rn}" for rn in public_rows_for_stats)
    cell = ws.cell(row=cur, column=col_idx)
    cell.value = f"=MEDIAN({refs})"
    cell.font = Font(bold=True)
    cell.fill = fill(MED_BLUE)
    cell.border = BORDER_THICK
    if col_idx in (7, 8):
        cell.number_format = "0.0\"x\""
    else:
        cell.number_format = "0.0%"
cur += 1

# Median for AI-for-science group
ai_bio_ids = ["recursion", "schrodinger", "relay-therapeutics", "roivant"]
ai_bio_rows_for_stats = []
for i, r in enumerate(public_company_rows):
    if r["id"] in ai_bio_ids:
        ai_bio_rows_for_stats.append(header_row + 1 + i)

ws.cell(row=cur, column=1).value = "Median (AI-for-science premium group)"
ws.cell(row=cur, column=1).font = Font(bold=True)
ws.cell(row=cur, column=1).border = BORDER_THIN
for col_idx in (7, 8, 9, 10):
    refs = ",".join(f"{chr(ord('A')+col_idx-1)}{rn}" for rn in ai_bio_rows_for_stats)
    cell = ws.cell(row=cur, column=col_idx)
    cell.value = f"=MEDIAN({refs})"
    cell.font = Font(bold=True)
    cell.fill = fill(LIGHT_BLUE)
    cell.border = BORDER_THICK
    if col_idx in (7, 8):
        cell.number_format = "0.0\"x\""
    else:
        cell.number_format = "0.0%"
cur += 2

# --- Implied valuation block ---
ws.cell(row=cur, column=1).value = "IMPLIED LUPINE VALUATION ($M)"
ws.cell(row=cur, column=1).font = Font(bold=True, color="FFFFFF")
ws.cell(row=cur, column=1).fill = fill(NAVY)
ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=11)
cur += 1

# Lupine ARR scenarios (from lupine_revenue_v2.csv)
lupine_revs = [
    ("FY27 base ARR", 4.20),
    ("FY28 base ARR", 16.15),
    ("FY29 base ARR", 35.00),
    ("FY30 base ARR", 62.40),
    ("FY32 base ARR", 133.30),
    ("FY30 bull ARR", 200.00),
    ("FY30 bear ARR", 22.00),
]

ws.cell(row=cur, column=1).value = "Lupine ARR scenario"
ws.cell(row=cur, column=2).value = "ARR ($M)"
ws.cell(row=cur, column=3).value = "Sim median (15x) EV"
ws.cell(row=cur, column=4).value = "AI-bio median (40x) EV"
for c in range(1, 5):
    cell = ws.cell(row=cur, column=c)
    cell.font = Font(bold=True)
    cell.fill = fill(LIGHT_BLUE)
    cell.border = BORDER_THICK
cur += 1

for label, arr in lupine_revs:
    ws.cell(row=cur, column=1).value = label
    ws.cell(row=cur, column=1).border = BORDER_THIN
    cell = ws.cell(row=cur, column=2)
    cell.value = arr
    cell.font = Font(color="0000FF")
    cell.fill = fill(LIGHT_GREY)
    cell.number_format = "$#,##0.0"
    cell.border = BORDER_THIN
    # Implied EV at simulation median
    cell = ws.cell(row=cur, column=3)
    cell.value = f"={chr(ord('A')+1)}{cur}*15"
    cell.number_format = "$#,##0"
    cell.border = BORDER_THIN
    # Implied EV at AI-bio median
    cell = ws.cell(row=cur, column=4)
    cell.value = f"={chr(ord('A')+1)}{cur}*40"
    cell.number_format = "$#,##0"
    cell.border = BORDER_THIN
    cur += 1

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}")
